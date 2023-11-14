import openpyxl
import sys
import base64

import jinja2
import pdfkit
from datetime import datetime, timedelta

from email.message import EmailMessage





    

# Get the right workbook and worksheet
wb = openpyxl.load_workbook('customers.xlsx')
sheet = wb['Transfer']

# Get the right row and column with data entry
last_row = sheet.max_row
last_col = sheet.max_column

# Convert the data entry to a proper dictionary, each key is a job number
# corresponsing value is the job details.
jobs = {}
for row in sheet.iter_rows(min_row=2, max_col=last_col, max_row=last_row, values_only=True):
    job_details = {sheet.cell(row=1, column=j).value: value for j, value in enumerate(row, start=1)}
    job_num = row[0]
    jobs[job_num] = job_details

# Filter the jobs that the invoice hasn't been send to the corresponding customer
jobs_to_send_invoices = {job_num: job_details for (job_num, job_details) in jobs.items() if job_details['Invoice Date'] == None}

for k, v in jobs_to_send_invoices.items():
    print(f"{k}: {v}")



for job_num, job_details in jobs_to_send_invoices.items():
    if not job_details["Invoice Date"]:
        billing_info = job_details["Billing Info"]
        billing_info_optional = job_details["Billing Info(optional)"]
        job_no = job_details["Job No"]
        invoice_date = datetime.today().strftime("%d %b %Y")
        due_date = (datetime.today() + timedelta(days=7)).strftime("%d %b %Y")
        rate = float(job_details["Rate"])
        hours = job_details["Hours"]
        removal_fee_total = rate * hours
        surcharge = float(job_details["Surcharge"])
        sub_total = float(removal_fee_total + surcharge)
        gst = float(sub_total * 0.1)
        total = float(sub_total + gst)
        
        content = {
            "billing_info": billing_info,
            "job_no": job_no,
            "invoice_date": invoice_date,
            "rate": f"$ {rate:.2f}",
            "hours": hours,
            "removal_fee_total": f"$ {removal_fee_total:.2f}",
            "surcharge": f"$ {surcharge:.2f}",
            "sub_total": f"$ {sub_total:.2f}",
            "gst": f"$ {gst:>8.2f}",
            "total": f"$ {total:>8.2f}",
            "due_date": due_date,
        }
        if billing_info_optional is not None:
            content["billing_info_optional"] = billing_info_optional

        template_loader = jinja2.FileSystemLoader("./")
        template_env = jinja2.Environment(loader=template_loader)

        html_template = "invoice.html"
        template = template_env.get_template(html_template)
        output_text = template.render(content)

        config = pdfkit.configuration(wkhtmltopdf="/usr/local/bin/wkhtmltopdf")
        output_pdf = f"{job_no}.pdf"
        pdfkit.from_string(output_text, output_pdf, configuration=config, css="invoice.css")